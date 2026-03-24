"""
ERP 발주저장 양식(260319).xls  Sheet1 기준으로
동일한 양식의 시트 3개를 생성
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import date, timedelta

today       = date.today()
today_plus5 = today + timedelta(days=5)

def excel_date(d):
    """Python date → Excel 시리얼 번호"""
    from datetime import date as _d
    origin = _d(1899, 12, 30)
    return (d - origin).days

def fmt_label(d):
    return f"{str(d.year)[2:]}{d.month:02d}{d.day:02d}"

DATE_VAL   = excel_date(today)         # 발주일 (Excel serial)
PLUS5_VAL  = excel_date(today_plus5)   # 발송/납기일 (Excel serial)
LABEL_TODAY = fmt_label(today)

# ── border 헬퍼 ──────────────────────────────────────────────
def S(style):
    return Side(style=style) if style else Side(style=None)

none  = Side(style=None)
thin  = Side(style='thin')
med   = Side(style='medium')
hair  = Side(style='hair')

def B(t=None, b=None, l=None, r=None):
    return Border(top=S(t), bottom=S(b), left=S(l), right=S(r))

YELLOW = PatternFill('solid', fgColor='FFFF00')
DATE_FMT = 'YY-M-D'

def apply(c, value='', bold=False, size=10, halign='center', valign='center',
          wrap=False, fill=None, bdr=None, numfmt=None):
    c.value = value
    c.font  = Font(name='Arial', bold=bold, size=size)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if fill:  c.fill   = fill
    if bdr:   c.border = bdr
    if numfmt: c.number_format = numfmt

def mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)

def build_sheet(ws):
    # ── 열 너비 (원본 256단위 / 256 = 문자 폭) ──────────────
    widths = [4.46, 15.53, 15.86, 3.80, 8.86, 8.86, 1.80, 12.20, 9.20, 8.86, 8.86, 8.86]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # ── 행 높이 (원본 20단위 / 20 = pt) ─────────────────────
    rh = {1:22.5, 2:13.5, 3:16.5, 4:16.5, 5:16.5,
          6:13.5, 7:16.5, 8:16.5, 9:16.5,
          10:13.5, 11:16.5, 12:13.5, 13:13.5,
          14:17.25, 15:17.25, 16:16.5,
          17:24.75, 18:24.75}
    for r in range(19, 81):
        rh[r] = 18.0
    for r, h in rh.items():
        ws.row_dimensions[r].height = h

    w = ws.cell  # shorthand

    # ━━ ROW 1 : PROFORMA INVOICE ━━━━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 1,1, 1,9)
    apply(w(1,1), 'PROFORMA INVOICE', bold=True, size=14,
          bdr=B(b='medium'))

    # ━━ ROW 2 : SHIPPER / NO & DATE ━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 2,1, 2,2)
    apply(w(2,1), 'SHIPPER', bold=True, bdr=B(t='medium', b='thin'))
    apply(w(2,3), '', bdr=B(r='thin'))
    mg(ws, 2,4, 2,6)
    apply(w(2,4), 'NO & DATE OF INVOICE', bold=True, bdr=B(b='thin'))

    # ━━ ROW 3 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    apply(w(3,1), 'SOLTRI CORPORATION', halign='left')
    apply(w(3,3), '', bdr=B(r='thin'))

    # ━━ ROW 4 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    apply(w(4,1), '133B, 1L, 704, GOJAN-DONG, NAMDONG-GU', halign='left')
    apply(w(4,3), '', bdr=B(r='thin'))

    # ━━ ROW 5 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    apply(w(5,1), 'INCHEON KOREA TEL : 82-32-761-2323', halign='left',
          bdr=B(b='thin'))
    apply(w(5,3), '', bdr=B(b='thin', r='thin'))
    for c in [4,5,6,7,8,9]:
        apply(w(5,c), '', bdr=B(b='thin'))

    # ━━ ROW 6 : 업체명 / 발주일 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 6,1, 6,2)
    apply(w(6,1), '업체명', bold=True, bdr=B(t='thin', b='thin'))
    apply(w(6,3), '', bdr=B(r='thin'))
    mg(ws, 6,4, 6,6)
    apply(w(6,4), '발주일(수정)', bold=True,
          bdr=B(t='thin', b='thin', l='thin', r='thin'))

    # ━━ ROW 7 : WIPRO / 발주일 값 ━━━━━━━━━━━━━━━━━━━━━━━━━━
    apply(w(7,2), 'WIPRO', bold=True)
    apply(w(7,3), '', bdr=B(r='thin'))
    apply(w(7,5), DATE_VAL, bold=True, fill=YELLOW, numfmt=DATE_FMT)

    # ━━ ROW 8 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    apply(w(8,3), '', bdr=B(r='thin'))

    # ━━ ROW 9 : 구분선 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    for c in range(1, 10):
        apply(w(9,c), '', bdr=B(b='thin'))

    # ━━ ROW 10 : NOTIFY / REMARKS ━━━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 10,1, 10,2)
    apply(w(10,1), 'NOTIFY', bold=True, bdr=B(t='thin', b='thin'))
    apply(w(10,3), '', bdr=B(r='thin'))
    mg(ws, 10,4, 10,6)
    apply(w(10,4), 'REMARKS', bold=True, bdr=B(b='thin'))

    # ━━ ROW 11 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    for c in [1,2,3]:
        apply(w(11,c), '', bdr=B(b='thin'))

    # ━━ ROW 12 : LOADING PORT ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 12,1, 12,2)
    apply(w(12,1), 'LOADING PORT', bold=True, bdr=B(t='thin', b='thin'))
    apply(w(12,3), '', bdr=B(t='thin', b='thin', l='thin', r='thin'))

    # ━━ ROW 13 : DESTINATION ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 13,1, 13,2)
    apply(w(13,1), 'DESTINATION', bold=True, bdr=B(t='thin', b='thin'))
    for c in range(3, 10):
        apply(w(13,c), '', bdr=B(b='thin'))

    # ━━ ROW 14 : VESSEL / 발송일자 / 안전재고 ━━━━━━━━━━━━━━━
    mg(ws, 14,1, 14,2)
    apply(w(14,1), 'VESSEL & VOY', bold=True, bdr=B(t='thin', b='thin'))
    apply(w(14,3), '', bdr=B(t='thin', b='thin', l='thin', r='thin'))
    mg(ws, 14,4, 14,6)
    apply(w(14,4), '발송일자', bold=True, bdr=B(t='thin', b='thin'))
    mg(ws, 14,7, 14,9)
    apply(w(14,7), f'안전재고 ({LABEL_TODAY})', bold=True, fill=YELLOW,
          bdr=B(t='thin', b='thin', l='thin'))

    # ━━ ROW 15 : 납기일자 / ETA ━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 15,1, 15,2)
    apply(w(15,1), '납기일자(수정)', bold=True, bdr=B(t='thin', b='thin'))
    apply(w(15,3), PLUS5_VAL, bold=True, fill=YELLOW,
          bdr=B(b='thin'), numfmt=DATE_FMT)
    mg(ws, 15,4, 15,6)
    apply(w(15,4), 'ETA', bold=True, bdr=B(t='thin', b='thin'))
    for c in range(7, 10):
        apply(w(15,c), '', bdr=B(b='thin'))

    # ━━ ROW 16 : DESCRIPTION ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    mg(ws, 16,1, 16,9)
    apply(w(16,1), 'DESCRIPTION : PLAIN SHAFT BEARING - HYDRAULIC SEALS',
          bold=True, halign='left', bdr=B(t='thin', b='thin'))

    # ━━ ROW 17 : 테이블 헤더 ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    apply(w(17,1), 'NO.',                    bdr=B(t='thin', b='thin', r='hair'))
    apply(w(17,2), '품번(19번부터인식)', bold=True, wrap=True,
          bdr=B(t='thin', b='thin', l='hair'))
    apply(w(17,3), '치수',          bold=True, bdr=B(t='thin', b='thin'))
    apply(w(17,4), '',              bdr=B(t='thin', b='thin'))
    apply(w(17,5), '품명/재질',     bold=True, bdr=B(t='thin', b='thin'))
    apply(w(17,6), '수량',          bold=True, halign='center',
          bdr=B(t='thin', b='thin', l='hair'))
    apply(w(17,7), '',              bdr=B(t='thin', b='thin', r='hair'))
    apply(w(17,8), 'U/PRICE (USD)(발송일자)', bold=True, wrap=True,
          bdr=B(t='thin', l='hair', r='hair'))
    apply(w(17,9), 'AMOUNT (USD)',  bdr=B(t='thin', b='thin'))

    # ━━ ROW 18 : H열 상단 보조선 ━━━━━━━━━━━━━━━━━━━━━━━━━━━
    apply(w(18,8), '', bdr=B(t='thin'))

    # ━━ ROW 19+ : 데이터 행 (원본 수량 그대로) ━━━━━━━━━━━━━━
    raw_data = [
        (1, '', '45 x 50 x 9.5',  'W', 'W/CN10', 380),
        (2, '', '130 x 135 x 15', 'W', 'W/CN10', 220),
        (3, '', '58 x 63 x 20',   'W', 'W/CN10', 240),
    ]
    for i, (no, pn, dim, mat, name, qty) in enumerate(raw_data):
        r = 19 + i
        apply(w(r,1), no,   bdr=B(t='hair', b='hair'))
        apply(w(r,2), pn,   bdr=B(t='hair', b='hair'))
        apply(w(r,3), dim,  halign='left', bdr=B(t='hair', b='hair'))
        apply(w(r,4), mat,  halign='left', bdr=B(t='hair', b='hair'))
        apply(w(r,5), name, halign='left', bdr=B(t='hair', b='hair'))
        apply(w(r,6), qty,  halign='right', bdr=B(t='hair', b='hair'))
        apply(w(r,7), '',   bdr=B(t='hair', b='hair'))
        apply(w(r,8), PLUS5_VAL, bold=True, fill=YELLOW,
              bdr=B(b='thin'), numfmt=DATE_FMT)
        apply(w(r,9), '',   bdr=B(t='hair', b='hair'))

    # 빈 데이터 행 (22~80)
    for r in range(22, 81):
        apply(w(r,8), '', bdr=B(b='thin'))

# ── 워크북 생성, 3개 시트 ─────────────────────────────────────
wb = openpyxl.Workbook()

sheet_names = ['Sheet1', 'Sheet2', 'Sheet3']
ws1 = wb.active
ws1.title = sheet_names[0]
build_sheet(ws1)

for name in sheet_names[1:]:
    ws = wb.create_sheet(title=name)
    build_sheet(ws)

output = rf"G:\내 드라이브\work\frequency\ERP_발주저장_{LABEL_TODAY}.xlsx"
wb.save(output)
print(f"저장 완료: {output}")
