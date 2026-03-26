# -*- coding: utf-8 -*-
"""
1) ERP 엑셀 3개 읽기 (라벨조회 + 완제품재고창고 + 가공작업조회)
2) data500.json 생성
3) data_wr.json 생성
4) index.html 재생성
"""
import openpyxl, json, re, os
from collections import Counter, defaultdict
from datetime import datetime, timedelta

# ── 경로 설정 (하드코딩 제거) ──────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WORK_DIR = os.path.dirname(SCRIPT_DIR)
ERP = os.path.join(WORK_DIR, "ERP_downloaded_data")
OUT = SCRIPT_DIR
CREDS_PATH = os.path.join(WORK_DIR, "keys", "gen-lang-client-0766779209-5009b1c068c6.json")
TODAY = datetime.now()
WEEK_AGO = TODAY - timedelta(days=7)

print(f"=== Soltri 선생산 리프레시 ({TODAY.strftime('%Y-%m-%d %H:%M')}) ===\n")

# ── graveyard.json 로드 (대시보드에서 '생산지시 중'으로 보낸 품목) ──
GRAVE_FILE = os.path.join(OUT, 'graveyard.json')
graveyard_keys = set()
try:
    if os.path.exists(GRAVE_FILE):
        with open(GRAVE_FILE, 'r', encoding='utf-8') as f:
            graveyard_data = json.load(f)
            graveyard_keys = set(graveyard_data.keys())
        print(f"[0] graveyard.json: {len(graveyard_keys)}개 품목 생산지시 중")
except Exception as e:
    print(f"[0] graveyard.json 로드 실패: {e}")

# ==============================
# 1. 라벨조회 (주문 빈도)
# ==============================
print("[1/4] 라벨조회.xlsx 읽는중...")
try:
    wb = openpyxl.load_workbook(f'{ERP}/라벨조회.xlsx', read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data = rows[2:]
except Exception as e:
    print(f"  → 라벨조회 로드 실패: {e}")
    data = []

freq = Counter()
qty_sum = defaultdict(int)
customers = defaultdict(lambda: defaultdict(int))
customer_freq = defaultdict(lambda: defaultdict(int))
pumok_map = defaultdict(str)

safe_stock_keys = set()  # 안전재고 발주 품목 수집 (7일 생산이력에서 제외용)
for r in data:
    order_no = str(r[1]) if r[1] else ''
    if '안전재고' in order_no:
        if r[4] and r[5]:
            safe_stock_keys.add((r[4], r[5]))

for r in data:
    chisu, jaejil, qty = r[4], r[5], r[8] if r[8] else 0
    customer = str(r[0]) if r[0] else '미상'
    pumok = r[6]
    order_no = str(r[1]) if r[1] else ''
    if customer == 'HYUNDAI CN':
        continue
    if '안전재고' in order_no:
        continue
    if chisu and jaejil:
        key = (chisu, jaejil)
        freq[key] += 1
        qty_sum[key] += qty
        customers[key][customer] += qty
        customer_freq[key][customer] += 1
        if pumok:
            pumok_map[key] = pumok

print(f"  → {len(data)}건, HYUNDAI CN 제외")

# ==============================
# 2. 완제품재고창고 (현재고)
# ==============================
print("[2/4] 완제품재고창고.xlsx 읽는중...")
rows2 = []
try:
    wb2 = openpyxl.load_workbook(f'{ERP}/완제품재고창고.xlsx', read_only=True)
    ws2 = wb2[wb2.sheetnames[0]]
    rows2 = list(ws2.iter_rows(values_only=True))
    wb2.close()
except Exception as e:
    print(f"  → 완제품재고창고 로드 실패: {e}")

def fmt(v):
    if v is None or v == '' or v == 0 or v == '0': return None
    try:
        f = float(v); return str(int(f)) if f == int(f) else str(f)
    except (ValueError, TypeError): return str(v)

stock = {}
for r in rows2[1:]:
    inner, outer, sz, height, jaejil = r[0], r[1], r[2], r[3], r[4]
    qty = r[7]
    if not inner or not outer or not jaejil: continue
    i, o, h = fmt(inner), fmt(outer), fmt(height)
    if not i or not o: continue
    chisu = f"{i}*{o}*{h}" if h and h != '0' else f"{i}*{o}"
    key = f"{jaejil}|{chisu}"
    try: q = int(float(str(qty).replace(',', ''))) if qty else 0
    except (ValueError, TypeError): q = 0
    stock[key] = stock.get(key, 0) + q

if rows2:
    print(f"  → {len(rows2)-1}건, {len(stock)}개 품목")

# ==============================
# 3. 가공작업조회 (7일 이내 생산이력)
# ==============================
print("[3/4] 가공작업조회.xlsx 읽는중...")
recent_production = set()
try:
    wb3 = openpyxl.load_workbook(f'{ERP}/가공작업조회.xlsx', read_only=True)
    ws3 = wb3[wb3.sheetnames[0]]
    rows3 = list(ws3.iter_rows(values_only=True))
    wb3.close()

    for r in rows3[2:]:
        start_date = r[4]
        chisu = r[6]
        jaejil = r[7]
        if not chisu or not jaejil or not start_date:
            continue
        if isinstance(start_date, datetime) and start_date >= WEEK_AGO:
            recent_production.add((str(chisu), str(jaejil)))

    print(f"  → {len(rows3)-2}건, 최근 7일 생산: {len(recent_production)}개")
except Exception as e:
    print(f"  → 가공작업조회 로드 실패: {e}")

# ==============================
# 4. 결과 조합 + 저장
# ==============================
print("[4/4] 분석 및 저장...")

results = []
for key, count in freq.items():
    total = qty_sum[key]
    if count >= 3 and total <= 1000:
        cust_detail = []
        for c, q in sorted(customers[key].items(), key=lambda x: -x[1]):
            cust_detail.append({'name': c, 'qty': q, 'freq': customer_freq[key][c]})
        stk_key = f"{key[1]}|{key[0]}"
        stk = stock.get(stk_key, -1)
        in_prod = (key[0], key[1]) in recent_production
        # 안전재고 발주로 생산 중인 품목은 in_production 제외
        # → 대시보드의 graveyard/completed 흐름이 관리
        if in_prod and key in safe_stock_keys:
            in_prod = False
        # graveyard(생산지시 중)에 있는 품목도 in_production 제외
        item_key = f"{key[0]}|{key[1]}"
        if in_prod and item_key in graveyard_keys:
            in_prod = False
        # 재고율 30% 초과면 생산 불필요
        if stk >= 0 and stk > total * 0.3:
            in_prod = True

        results.append({
            'chisu': key[0], 'jaejil': key[1], 'pumok': pumok_map[key],
            'freq': count, 'total_qty': total, 'avg_qty': round(total/max(count, 1), 1),
            'customers': cust_detail, 'cust_count': len(cust_detail),
            'stock': stk, 'in_production': in_prod
        })

# 점수
for r in results:
    stk = r['stock']
    stock_ratio = (stk / max(r['total_qty'], 1)) if stk >= 0 else 0.5
    freq_score = (r['freq'] / 20) * 50
    if stock_ratio <= 0.05: urgency = 50
    elif stock_ratio <= 0.10: urgency = 40
    elif stock_ratio <= 0.20: urgency = 30
    elif stock_ratio <= 0.50: urgency = 15
    else: urgency = 0
    r['score'] = round(freq_score + urgency, 1)
    r['stock_ratio'] = round(stock_ratio * 100, 1)

results.sort(key=lambda x: x['score'], reverse=True)

# summary
js = defaultdict(lambda: {'count':0,'qty':0,'items':0})
ps = defaultdict(lambda: {'count':0,'qty':0,'items':0})
cs = defaultdict(lambda: {'count':0,'qty':0,'items':0})
for r in results:
    js[r['jaejil']]['items']+=1; js[r['jaejil']]['count']+=r['freq']; js[r['jaejil']]['qty']+=r['total_qty']
    ps[r['pumok']]['items']+=1; ps[r['pumok']]['count']+=r['freq']; ps[r['pumok']]['qty']+=r['total_qty']
    for c in r['customers']:
        cs[c['name']]['items']+=1; cs[c['name']]['count']+=c['freq']; cs[c['name']]['qty']+=c['qty']

output = {
    'results': results,
    'jaejil_summary': dict(js), 'pumok_summary': dict(ps), 'cust_summary': dict(cs),
    'total_items': len(results),
    'analysis_date': TODAY.strftime('%Y-%m-%d'),
    'exclude_customers': ['HYUNDAI CN'],
    'recent_production_count': len(recent_production)
}

with open(f'{OUT}/data500.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

# W/R 핵심 → data_wr.json
items = [r for r in results if r['pumok'] == 'W/R' and r['freq'] >= 5]
items.sort(key=lambda x: x['score'], reverse=True)
for r in items:
    r['key'] = f"{r['chisu']}|{r['jaejil']}"

with open(f'{OUT}/data_wr.json', 'w', encoding='utf-8') as f:
    json.dump(items, f, ensure_ascii=False, separators=(',', ':'))

# data_wr.js (파일 직접 열기용)
with open(f'{OUT}/data_wr.js', 'w', encoding='utf-8') as f:
    f.write('var DATA_WR = ')
    json.dump(items, f, ensure_ascii=False, separators=(',', ':'))
    f.write(';')

active = [r for r in items if not r['in_production']]
in_prod = [r for r in items if r['in_production']]

print(f"\n  data500.json: {len(results)}개 품목")
print(f"  data_wr.json: {len(items)}개 (선생산 {len(active)} + 생산중 {len(in_prod)})")

# ==============================
# 5. 발주조회 (최근 발주일)
# ==============================
print("[5/6] 발주조회.xlsx 읽는중...")
try:
    wb4 = openpyxl.load_workbook(f'{ERP}/발주조회.xlsx', read_only=True)
    ws4 = wb4[wb4.sheetnames[0]]
    rows4 = list(ws4.iter_rows(values_only=True))
    wb4.close()

    # W/R 핵심 치수 파싱 (순서 무관 매칭용)
    wr_nums = {}  # frozenset(parts)|jaejil -> key
    for r in items:
        parts = r['chisu'].split('*')
        wr_nums[frozenset(parts) | {r['jaejil']}] = r['key']

    # 업체치수 파싱: "WEAR RING_PW D110D105W15A" → ['110','105','15A']
    def parse_balju_chisu(s):
        m = re.search(r'D([\d.]+)D([\d.]+)W([\w.]+)', str(s))
        return [m.group(1), m.group(2), m.group(3)] if m else None

    latest_balju = {}  # key -> latest date string
    for r in rows4[2:]:
        if not r[15] or not r[17] or r[18] != 'W/R': continue
        if '안전재고' not in str(r[5] or ''): continue  # 안전재고 발주만 인식
        jaejil = str(r[17])
        parts = parse_balju_chisu(r[15])
        if not parts: continue

        lookup = frozenset(parts) | {jaejil}
        if lookup in wr_nums:
            key = wr_nums[lookup]
            order_date = r[6]
            if isinstance(order_date, datetime):
                ds = order_date.strftime('%Y-%m-%d')
            else:
                ds = str(order_date)[:10]
            if key not in latest_balju or ds > latest_balju[key]:
                latest_balju[key] = ds

    # items에 최근 발주일 추가
    for r in items:
        r['latest_balju'] = latest_balju.get(r['key'], '')

    # data_wr.json + data_wr.js 재저장
    with open(f'{OUT}/data_wr.json', 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, separators=(',', ':'))
    with open(f'{OUT}/data_wr.js', 'w', encoding='utf-8') as f:
        f.write('var DATA_WR = ')
        json.dump(items, f, ensure_ascii=False, separators=(',', ':'))
        f.write(';')

    print(f"  → {len(rows4)-2}건, 매칭 {len(latest_balju)}개")
except Exception as e:
    print(f"  → 발주조회 로드 실패: {e}")

# ==============================
# 6. Google Sheets 업로드
# ==============================
print("\n[6/6] Google Sheets 업로드...")
try:
    import gspread
    from google.oauth2.service_account import Credentials

    SHEET_ID = "1zkyYFiX5MGkGj7cnjhpiPvlzHLwN0CZKMFnjl9tTzqA"
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

    creds = Credentials.from_service_account_file(CREDS_PATH, scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SHEET_ID)

    # 탭 초기화
    try:
        old_ws = sh.worksheet("선생산리스트")
        sh.del_worksheet(old_ws)
    except gspread.exceptions.WorksheetNotFound:
        pass

    ws = sh.add_worksheet(title="선생산리스트", rows=len(items)+5, cols=15)

    # 헤더 + 데이터
    header = ['치수','재질','품목','주문횟수','총수량','평균/회','현재고','재고율(%)','생산중','점수','거래처','최근발주일']
    sheet_rows = [header]
    for r in items:
        custs = ', '.join([c['name'] for c in r['customers']])
        sheet_rows.append([
            r['chisu'], r['jaejil'], r['pumok'],
            r['freq'], r['total_qty'], r['avg_qty'],
            r['stock'], r['stock_ratio'],
            'Y' if r['in_production'] else 'N',
            r['score'], custs, r.get('latest_balju', '')
        ])

    ws.update(values=sheet_rows, range_name=f'A1:L{len(sheet_rows)}')
    print(f"  → 선생산리스트 업로드 완료 ({len(items)}행)")

    # 주문이력 탭
    wr_keys = set(r['key'] for r in items)
    history = []
    for r in data:
        customer = str(r[0]) if r[0] else ''
        order_date = r[2]
        chisu = r[4]
        jaejil = r[5]
        qty = r[8] if r[8] else 0
        order_no = str(r[1]) if r[1] else ''
        if not chisu or not jaejil: continue
        if customer == 'HYUNDAI CN': continue
        if '안전재고' in order_no: continue
        key = f"{chisu}|{jaejil}"
        if key not in wr_keys: continue
        date_str = order_date.strftime('%Y-%m-%d') if isinstance(order_date, datetime) else str(order_date)
        history.append([str(chisu), str(jaejil), date_str, customer, int(qty)])

    # 발주조회 데이터도 이력에 추가 (타입=발주)
    for r4 in rows4[2:]:
        if not r4[15] or not r4[17] or r4[18] != 'W/R': continue
        if '안전재고' in str(r4[5] or ''): continue
        jaejil4 = str(r4[17])
        parts4 = parse_balju_chisu(r4[15])
        if not parts4: continue
        lookup4 = frozenset(parts4) | {jaejil4}
        if lookup4 not in wr_nums: continue
        wkey = wr_nums[lookup4]
        chisu4 = wkey.split('|')[0]
        od = r4[6]
        ds4 = od.strftime('%Y-%m-%d') if isinstance(od, datetime) else str(od)[:10]
        cust4 = str(r4[4]) if r4[4] else ''
        qty4 = int(r4[20]) if r4[20] else 0
        history.append([chisu4, jaejil4, ds4, cust4, qty4, '발주'])

    # 라벨조회 건에 타입 추가
    for h in history:
        if len(h) == 5:
            h.append('주문')

    history.sort(key=lambda x: (x[0], x[1], x[2]))

    try:
        old_hist = sh.worksheet("주문이력")
        sh.del_worksheet(old_hist)
    except gspread.exceptions.WorksheetNotFound:
        pass

    ws_hist = sh.add_worksheet(title="주문이력", rows=len(history)+5, cols=6)
    hist_header = ['치수','재질','주문일자','거래처','수량','타입']
    hist_rows = [hist_header] + history
    ws_hist.update(values=hist_rows, range_name=f'A1:F{len(hist_rows)}')
    print(f"  → 주문이력 업로드 완료 ({len(history)}행)")

except Exception as e:
    print(f"  → Google Sheets 업로드 실패: {e}")
    print(f"  → data_wr.json은 정상 저장됨 (로컬 사용 가능)")

# ==============================
# 7. 변경 감지 + Outlook 이메일 알림
# ==============================
print("\n[7/7] 변경 감지 및 이메일 알림...")
try:
    PREV_FILE = os.path.join(OUT, 'data_wr_prev.json')
    prev_items = {}
    if os.path.exists(PREV_FILE):
        with open(PREV_FILE, 'r', encoding='utf-8') as f:
            for r in json.load(f):
                prev_items[r['key']] = r

    curr_items = {r['key']: r for r in items}

    # 신규 품목 (이전에 없던 것)
    new_keys = [k for k in curr_items if k not in prev_items]
    # 제거된 품목 (이전에 있었는데 없어진 것)
    removed_keys = [k for k in prev_items if k not in curr_items]
    # 재고 소진 (재고율 10% 이하로 떨어진 품목)
    stock_alert = []
    for k, r in curr_items.items():
        if k in prev_items:
            prev_ratio = prev_items[k].get('stock_ratio', 0)
            if prev_ratio > 10 and r['stock_ratio'] <= 10:
                stock_alert.append(r)

    # 선생산 필요 목록 (in_production=False, 재고율 30% 이하)
    active_list = [r for r in items if not r['in_production'] and r.get('stock_ratio', 100) <= 30]

    has_changes = new_keys or removed_keys or stock_alert
    print(f"  신규: {len(new_keys)}개, 제거: {len(removed_keys)}개, 재고소진경고: {len(stock_alert)}개")

    # 현재 데이터를 prev로 저장 (다음 비교용)
    import shutil
    shutil.copy2(f'{OUT}/data_wr.json', PREV_FILE)

    if has_changes or active_list:
        import win32com.client as win32

        # 이메일 본문 구성
        body = f'<html><body style="font-family:Malgun Gothic,sans-serif;font-size:13px;color:#1a1a1a">'
        body += f'<h2 style="color:#1e3a5f;border-bottom:2px solid #3b82f6;padding-bottom:6px">Soltri 선생산 리포트 ({TODAY.strftime("%Y-%m-%d %H:%M")})</h2>'

        if new_keys:
            body += f'<h3 style="color:#22c55e">🟢 신규 품목 ({len(new_keys)}개)</h3><table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse;font-size:12px">'
            body += '<tr style="background:#e8f5e9"><th>치수</th><th>재질</th><th>주문횟수</th><th>총량</th><th>현재고</th></tr>'
            for k in new_keys[:20]:
                r = curr_items[k]
                body += f'<tr><td>{r["chisu"]}</td><td>{r["jaejil"]}</td><td>{r["freq"]}회</td><td>{r["total_qty"]}</td><td>{r["stock"]}</td></tr>'
            body += '</table><br>'

        if removed_keys:
            body += f'<h3 style="color:#ef4444">🔴 제거된 품목 ({len(removed_keys)}개)</h3><table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse;font-size:12px">'
            body += '<tr style="background:#fce4ec"><th>치수</th><th>재질</th></tr>'
            for k in removed_keys[:20]:
                parts = k.split('|')
                body += f'<tr><td>{parts[0]}</td><td>{parts[1] if len(parts)>1 else ""}</td></tr>'
            body += '</table><br>'

        if stock_alert:
            body += f'<h3 style="color:#f59e0b">⚠️ 재고 소진 경고 ({len(stock_alert)}개)</h3><table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse;font-size:12px">'
            body += '<tr style="background:#fff8e1"><th>치수</th><th>재질</th><th>현재고</th><th>재고율</th></tr>'
            for r in stock_alert[:20]:
                body += f'<tr><td>{r["chisu"]}</td><td>{r["jaejil"]}</td><td>{r["stock"]}</td><td>{r["stock_ratio"]}%</td></tr>'
            body += '</table><br>'

        # 선생산 필요 TOP 10
        top10 = sorted(active_list, key=lambda x: x['score'], reverse=True)[:10]
        if top10:
            body += f'<h3 style="color:#3b82f6">📋 선생산 필요 TOP 10 (점수순)</h3><table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse;font-size:12px">'
            body += '<tr style="background:#e3f2fd"><th>#</th><th>치수</th><th>재질</th><th>주문</th><th>총량</th><th>현재고</th><th>재고율</th><th>점수</th></tr>'
            for i, r in enumerate(top10, 1):
                bg = '#fff5f5' if r['stock_ratio'] <= 10 else '#ffffff'
                body += f'<tr style="background:{bg}"><td>{i}</td><td>{r["chisu"]}</td><td>{r["jaejil"]}</td><td>{r["freq"]}회</td><td>{r["total_qty"]}</td><td>{r["stock"]}</td><td>{r["stock_ratio"]}%</td><td>{r["score"]}</td></tr>'
            body += '</table><br>'

        body += f'<p style="color:#94a3b8;font-size:11px">총 {len(items)}개 품목 | 선생산 필요 {len(active_list)}개 | <a href="https://nuckythompson28-code.github.io/frequency/안전재고발주.html">대시보드 열기</a></p>'
        body += '</body></html>'

        # 제목
        subject = f'[선생산] '
        changes = []
        if new_keys: changes.append(f'신규 {len(new_keys)}')
        if removed_keys: changes.append(f'제거 {len(removed_keys)}')
        if stock_alert: changes.append(f'재고소진 {len(stock_alert)}')
        if changes:
            subject += ', '.join(changes) + f' | 선생산 필요 {len(active_list)}개'
        else:
            subject += f'선생산 필요 {len(active_list)}개 ({TODAY.strftime("%m/%d")})'

        # Outlook 발송
        outlook = win32.Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.To = 'steven@soltri.com; aileen@soltri.com; onestar@soltri.com'
        mail.Subject = subject
        mail.HTMLBody = body
        mail.Send()
        print(f"  → 이메일 발송 완료: {subject}")

    else:
        print("  → 변경 없음, 이메일 미발송")

except Exception as e:
    print(f"  → 이메일 알림 실패: {e}")

print(f"\n완료!")
